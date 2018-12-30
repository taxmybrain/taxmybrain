import string
import decimal
from flask import (
    Blueprint, render_template, request, redirect, url_for, session,
    current_app, jsonify, flash,
)
from werkzeug.utils import secure_filename
import requests
from requests_oauthlib import OAuth2Session
import urllib
from openpyxl import load_workbook
from . import main


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in \
        current_app.config['ALLOWED_EXTENSIONS']


def buildheaders(authcode='', govtest=''):
    headers = {
        'User-Agent': current_app.config['USER_AGENT'],
        'Accept': current_app.config['ACCEPT'],
        'Content-Type': 'application/json',
        'Authorization': 'Bearer {authcode}'.format(authcode=authcode),
        'Gov-Test-Scenario': govtest, # For testing in sandbox
    }
    return headers


def buildurl(endpoint, vrn='', period='', parameters=[], api='test-api'):
    """Builds endpoint URLS, including parameters.
    For production api='api'
    """
    if parameters != []:
        question = '?'
    else:
        question = ''
    escaped_period = urllib.parse.quote(period)
    # html escape in case there is a '#'
    baseurl = 'https://{api}.service.hmrc.gov.uk/{service}{question}'.format(
        api=api,
        service=current_app.config[endpoint].format(vatno=vrn,
            period=escaped_period),
        question=question,
    )
    # note below that parameters is a list of tuples, not a dict.
    # This is to ensure order of parameters is consistant for testing.
    return baseurl + urllib.parse.urlencode(parameters)


@main.route("/workingdemostep1")
def workingdemostep1():
    return render_template('workingdemostep1.html')


@main.route("/workingdemostep2")
def workingdemostep2():
    userdata = create_user().json()
    # For the prototype, store the test user's vat number in a session cookie
    # to avoid having to type it in.
    session['user_vrn'] = userdata['vrn']
    return render_template('workingdemostep2.html', userdata=userdata)
    

@main.route("/workingdemostep3")
def workingdemostep3():
    obls = obligations()
    return render_template('workingdemostep3.html', obligations=obls)


@main.route("/")
@main.route("/index")
def index():
    return render_template('index.html')


@main.route("/login")
@main.route("/login/<returnpage>")
def login(returnpage='main.index'):
    """returnpage is the page to return to after HMRC calls the callback
       endpoint
    """
    session['returnpage'] = returnpage
    hmrc = OAuth2Session(current_app.config['CLIENT_ID'],
        redirect_uri=current_app.config['REDIRECT_URI_TEST'],
        scope=current_app.config['SCOPES'],
    )
    hmrc.headers=buildheaders()
    authorization_url, state = hmrc.authorization_url(buildurl(
        'AUTHORIZATION_BASE_URL', api='test-api'))
    # State is used to prevent CSRF, keep this for later.
    session['oauth_state'] = state
    return redirect(authorization_url)


@main.route("/auth-redirect")
def callback():
    """Callback from HMRC website once permission has been granted.
        When fetching the token https should be used and authorization_response
        requires it.
        For this protype we will bypass this and get the code ourselves from
        the return url.
    """
    hmrc = OAuth2Session(current_app.config['CLIENT_ID'],
        redirect_uri=current_app.config['REDIRECT_URI_TEST'],
        scope=current_app.config['SCOPES'],
        state=session['oauth_state'],
    )
    # Use the following in the prototype
    token = hmrc.fetch_token(buildurl('TOKEN_URL', api='test-api'),
        client_secret=current_app.config['CLIENT_SECRET'],
        code=request.args.get('code'),
    )
    # Use the following in production with https
    # token = hmrc.fetch_token(buildurl('TOKEN_URL', api='test-api'),
    #     client_secret=current_app.config['CLIENT_SECRET'],
    #     authorization_response=request.url,
    #     )
    
    # Store the access token in the session cookie
    session['access_token'] = token['access_token']
    return redirect(url_for(session['returnpage']))


@main.route('/obligations')
@main.route('/obligations/<fromdate>/<todate>')
def get_obligations(fromdate=None, todate=None):
    return jsonify(obligations(fromdate, todate))


def obligations(fromdate=None, todate=None):
    if fromdate is None:
        fromdate = '2017-01-01'
    if todate is None:
        todate = '2017-12-01'
    # paramters is list of tuples. urlencode can also accept a dict, but the
    # unordered nature of a dict meant that parameters were getting jumbled in
    # the url, causing testing problems.
    parameters = [('from', fromdate), ('to', todate)]
    # TODO need to add status??
    # parameters['status'] = 'O'
    obl = buildurl('URL_GET_OBLIGATIONS', vrn=session['user_vrn'],
        parameters=parameters, api='test-api')
    r = requests.get(obl, headers=buildheaders(session['access_token']))
    print(r.json())
    if r.json().get('code'):
        flash(r.json()['code'])
    return r.json()


@main.route('/workingdemostep4/<fromdate>/<todate>/<period>')
def vatdetails(fromdate, todate, period):    
    return render_template('workingdemostep4.html', fromdate=fromdate,
        todate=todate, period=period)


@main.route('/submit')
def submit_vat():
    submit = buildurl('URL_SUBMIT_VAT', vrn=session['user_vrn'], api='test-api')
    data = {"periodKey": "#001",
            "vatDueSales": 81293.59,#100.00,
            "vatDueAcquisitions": 0,#100.00,
            "totalVatDue": 81293.59,#200.00,
            "vatReclaimedCurrPeriod": 73510.64,#100.00,
            "netVatDue": 7782.95,#100.00,
            "totalValueSalesExVAT": 406468.16,#500,
            "totalValuePurchasesExVAT": 375429.39,#500,
            "totalValueGoodsSuppliedExVAT": '0',#500,
            "totalAcquisitionsExVAT": '0',#500,
            "finalised": True}
    r = requests.post(submit, headers=buildheaders(session['access_token']),
        json=data)
    return jsonify(r.json())


@main.route('/return')
@main.route('/return/<period>')
def vat_return(period=None):
    if period is None:
        period = session['periodkey']
    ret = buildurl('URL_GET_RETURNS', vrn=session['user_vrn'], period=period,
        api='test-api')
    r = requests.get(ret, headers=buildheaders(session['access_token']))
    return jsonify(r.json())


@main.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit an empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            print([thing for thing in request.form.keys()])
            session['periodkey'] = request.form['periodkey']
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            celldata = ws['C1':'C9']
            data = {cell[0].coordinate:cell[0].value for cell in celldata}
            wb.close()
            session['vat_data'] = data
            print(data)
            return redirect('showvatcalcs')
    return '''
    <!doctype html>
    <title>Upload new File</title>
    <h1>Upload new File</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=file>
      <input type=submit value=Upload>
    </form>
    '''


@main.route('/showvatcalcs')
def showvatcalcs():
    data = session['vat_data']
    return render_template('workingdemostep5.html', data=data)


@main.route('/submitvat', methods=['GET', 'POST'])
def submit_vat_return():
    dog = request.form['vatDueSales']
    if request.method == 'POST':
        submit = buildurl('URL_SUBMIT_VAT', vrn=session['user_vrn'],
            api='test-api')
        data = {
            "periodKey": session['periodkey'],
            "vatDueSales": request.form['vatDueSales'],
            "vatDueAcquisitions": request.form['vatDueAcquisitions'],
            "totalVatDue": request.form['totalVatDue'],
            "vatReclaimedCurrPeriod": request.form['vatReclaimedCurrPeriod'],
            "netVatDue": request.form['netVatDue'],
            "totalValueSalesExVAT": request.form['totalValueSalesExVAT'],
            "totalValuePurchasesExVAT":
                request.form['totalValuePurchasesExVAT'],
            "totalValueGoodsSuppliedExVAT":
                request.form['totalValueGoodsSuppliedExVAT'],
            "totalAcquisitionsExVAT": request.form['totalAcquisitionsExVAT'],
            "finalised": True
        }
        print(data)
        print(checkvatvalues(data))
        r = requests.post(submit, headers=buildheaders(
            session['access_token']), json=data)
        return jsonify(r.json())


def checkvatvalues(vat_data):
    # periodkey must be 4 alphanumeric characters and occasionally a # symbol
    periodkey_ok = periodkey_valid(vat_data['periodKey'])
    # vatDueSales must be between -9999999999999.99 and 9999999999999.99
    # (2 decimal places)
    vatduesales_ok = number_valid(vat_data['vatDueSales'])
    # vatDueAcquisitions must be between -9999999999999.99 and 9999999999999.99
    # (2 decimal places)
    vatdueacquisitions_ok = number_valid(vat_data['vatDueAcquisitions'])
    # totalVatDue must be between -9999999999999.99 and 9999999999999.99
    # (2 decimal places)
    totalvatdue_ok = number_valid(vat_data['totalVatDue'])
    # vatReclaimedCurrPeriod must be between -9999999999999.99 and
    # 9999999999999.99 (2 decimal places)
    vatreclaimedcurrperiod_ok = number_valid(vat_data['vatReclaimedCurrPeriod'])
    # netVatDue must be between 0.00 and 99999999999.99 (2 decimal places)
    netvatdue_ok = number_valid(vat_data['netVatDue'], lowerbound='0.00')
    # totalValueSalesExVAT must be between -9999999999999 and 9999999999999
    # (integer)
    totalvaluesalesexvat_ok = number_valid(vat_data['totalValueSalesExVAT'],
        dec_places=-1)
    # totalValuePurchasesExVAT must be between -9999999999999 and 9999999999999
    # (integer)
    totalvaluepurchasesexvat_ok = number_valid(
        vat_data['totalValuePurchasesExVAT'], dec_places=-1)
    # totalValueGoodsSuppliedExVAT must be between -9999999999999 and
    # 9999999999999 (integer)
    totalvaluegoodssuppliedexvat_ok = number_valid(
        vat_data['totalValueGoodsSuppliedExVAT'], dec_places=-1)
    # totalAcquisitionsExVAT must be between -9999999999999 and 9999999999999
    # (integer)
    totalacquisitionsdexvat_ok = number_valid(
        vat_data['totalAcquisitionsExVAT'], dec_places=-1)
    return (periodkey_ok, vatduesales_ok, vatdueacquisitions_ok, totalvatdue_ok,
        vatreclaimedcurrperiod_ok, netvatdue_ok, totalvaluesalesexvat_ok,
        totalvaluepurchasesexvat_ok, totalvaluegoodssuppliedexvat_ok,
        totalacquisitionsdexvat_ok,
    )


def periodkey_valid(periodkey):
    # periodkey must be a string of 4 alphanumeric characters and occasionally
    # a '#' symbol
    if not isinstance(periodkey, str):
        return False
    if len(periodkey) != 4:
        return False
    validset = set(string.ascii_letters + string.digits + '#')
    periodkey_valid = all([l in validset for l in periodkey])
    return periodkey_valid


def number_valid(value, lowerbound='-9999999999999.99',
    upperbound='9999999999999.99', dec_places=2):
    # value is passed as a string
    # if value is an integer, dec_places == -1
    lowerbound = decimal.Decimal(lowerbound)
    upperbound = decimal.Decimal(upperbound)
    num_dec_places = value[::-1].find('.')
    if num_dec_places != dec_places:
        return False
    inrange = lowerbound <= decimal.Decimal(value) <= upperbound
    return inrange


def int_nodp_valid(value):
    # value is passed as a string
    # Must represent an integer between -9999999999999 and 9999999999999
    num_dec_places = value[::-1].find('.')
    valid_dec_places = -1
    if num_dec_places != valid_dec_places:
        return False
    inrange = -9999999999999 <= int(value) <= 9999999999999.99
    return inrange


@main.route('/liabilities')
@main.route('/liabilities/<fromdate>/<todate>')
def liabilities(fromdate=None, todate=None):
    if fromdate is None:
        fromdate = '2017-01-02'
    if todate is None:
        todate = '2017-02-02'
    parameters = {}
    parameters['from'] = fromdate
    parameters['to'] = todate
    liab = buildurl('URL_GET_LIABILITIES', vrn=session['user_vrn'],
        parameters=parameters, api='test-api')
    r = requests.get(liab, headers=buildheaders(session['access_token'],
        govtest='SINGLE_LIABILITY'))
    return jsonify(r.json())


@main.route('/payments')
@main.route('/payments/<fromdate>/<todate>')
def payments(fromdate=None, todate=None):
    if fromdate is None:
        fromdate = '2017-01-02'
    if todate is None:
        todate = '2017-02-02'
    parameters = {}
    parameters['from'] = fromdate
    parameters['to'] = todate
    pay = buildurl('URL_GET_PAYMENTS', vrn=session['user_vrn'],
        parameters=parameters, api='test-api')
    r = requests.get(pay, headers=buildheaders(session['access_token'],
        govtest='SINGLE_PAYMENT'))
    return jsonify(r.json())


def create_user():
    r = requests.post(buildurl('URL_CREATE_TEST_USER', api='test-api'),
        headers=buildheaders(current_app.config['SERVER_TOKEN']),
        json={"serviceNames": ["mtd-vat",]})
    return r


@main.route('/debug')
def debug():
    return jsonify(
        {'uservrn': session['user_vrn'],
         'accesstoken': session['access_token'],
          'periodkey': session['periodkey'],
        }
    )
